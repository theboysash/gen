import asyncio
import aiohttp
import csv
import os
import time
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
API_KEY = "AIzaSyDt-QOkiAg8fLRi1N4eGk2GTGGHp26pByk"

MARKETS = {
    "Orlando": [
        "Winter Park FL",
        "College Park Orlando FL",
        "Dr Phillips Orlando FL",
        "Lake Nona Orlando FL",
        "Altamonte Springs FL",
    ],
    "Tampa": [
        "South Tampa FL",
        "Seminole Heights Tampa FL",
        "Carrollwood Tampa FL",
        "Wesley Chapel FL",
        "Westshore Tampa FL",
    ],
    "St. Petersburg": [
        "Downtown St Petersburg FL",
        "Old Northeast St Petersburg FL",
        "Kenwood St Petersburg FL",
        "Pinellas Park FL",
        "Gulfport FL",
    ],
    "Clearwater": [
        "Clearwater Beach FL",
        "Countryside Clearwater FL",
        "Safety Harbor FL",
        "Dunedin FL",
        "Palm Harbor FL",
    ],
    "Lakeland": [
        "Downtown Lakeland FL",
        "South Lakeland FL",
        "Lakeland Highlands FL",
        "Winter Haven FL",
        "Auburndale FL",
    ],
    "Kissimmee": [
        "Downtown Kissimmee FL",
        "Poinciana FL",
        "Celebration FL",
        "St Cloud FL",
        "BVL Kissimmee FL",
    ],
    "Jacksonville": [
        "Mandarin Jacksonville FL",
        "Southside Jacksonville FL",
        "San Marco Jacksonville FL",
        "Riverside Jacksonville FL",
        "Fleming Island FL",
    ],
    "Fort Lauderdale": [
        "Coral Springs FL",
        "Plantation FL",
        "Davie FL",
        "Pompano Beach FL",
        "Wilton Manors FL",
    ],
    "West Palm Beach": [
        "Palm Beach Gardens FL",
        "Boynton Beach FL",
        "Boca Raton FL",
        "Delray Beach FL",
        "Wellington FL",
    ],
}

INDUSTRIES = [
    "Electricians",
    "Plumbers",
    "HVAC companies",
    "Roofers",
    "Remodelers",
    "Cleaning companies",
    "Pest control",
    "Landscapers",
    "Pool service companies",
    "Dentists",
    "Med spas",
    "Chiropractors",
    "Accountants",
    "Mortgage brokers",
    "Real estate agents",
    "Property managers",
    "Personal injury attorneys",
]

CONCURRENCY = 5  # parallel requests at a time


# ── Async API call ────────────────────────────────────────

async def search_places_async(
    session: aiohttp.ClientSession,
    semaphore: asyncio.Semaphore,
    query: str,
    city: str,
    area: str,
    industry: str,
    progress: dict,
) -> list:
    url = "https://places.googleapis.com/v1/places:searchText"
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": API_KEY,
        "X-Goog-FieldMask": (
            "places.id,places.displayName,places.formattedAddress,"
            "places.nationalPhoneNumber,places.websiteUri,"
            "places.rating,places.userRatingCount"
        ),
    }

    results = []

    async with semaphore:
        try:
            body = {"textQuery": query, "pageSize": 20}
            async with session.post(url, headers=headers, json=body, timeout=aiohttp.ClientTimeout(total=15)) as res:
                data = await res.json()

                if "error" in data:
                    print(f"  API error for '{query}': {data['error'].get('message', '')}")
                    return []

                places = data.get("places", [])
                for p in places:
                    results.append({
                        "city":         city,
                        "area":         area,
                        "industry":     industry,
                        "place_id":     p.get("id", ""),
                        "name":         p.get("displayName", {}).get("text", "Unknown"),
                        "address":      p.get("formattedAddress", ""),
                        "phone":        p.get("nationalPhoneNumber", ""),
                        "website":      p.get("websiteUri", ""),
                        "has_website":  "Yes" if p.get("websiteUri") else "No",
                        "rating":       p.get("rating", ""),
                        "review_count": p.get("userRatingCount", 0),
                    })

        except Exception as e:
            print(f"  Error for '{query}': {e}")

    progress["done"] += 1
    pct = round(progress["done"] / progress["total"] * 100)
    print(f"  [{progress['done']}/{progress['total']} — {pct}%] {query} → {len(results)} results")

    return results


# ── Main Scanner ──────────────────────────────────────────

async def run_scan_async() -> list:
    # Build all tasks
    tasks_meta = []
    for city, areas in MARKETS.items():
        for area in areas:
            for industry in INDUSTRIES:
                tasks_meta.append((city, area, industry))

    progress = {"done": 0, "total": len(tasks_meta)}
    print(f"Total queries: {progress['total']}")
    print(f"Running {CONCURRENCY} in parallel\n")

    semaphore = asyncio.Semaphore(CONCURRENCY)
    all_leads = []
    seen_ids  = set()

    connector = aiohttp.TCPConnector(limit=20)
    async with aiohttp.ClientSession(connector=connector) as session:
        tasks = [
            search_places_async(session, semaphore, f"{ind} in {area}", city, area, ind, progress)
            for city, area, ind in tasks_meta
        ]
        results = await asyncio.gather(*tasks)

    for batch in results:
        for lead in batch:
            pid = lead["place_id"]
            if pid and pid not in seen_ids:
                seen_ids.add(pid)
                all_leads.append(lead)

    print(f"\nTotal unique businesses: {len(all_leads)}")
    return all_leads


# ── Summary Stats ─────────────────────────────────────────

def build_summary(leads: list) -> dict:
    grouped = defaultdict(lambda: defaultdict(list))
    for lead in leads:
        grouped[lead["city"]][lead["industry"]].append(lead)

    summary = {}
    for city, industries in grouped.items():
        summary[city] = {}
        for industry, items in industries.items():
            total      = len(items)
            no_website = sum(1 for x in items if x["has_website"] == "No")
            ratings    = [float(x["rating"]) for x in items if x["rating"] != ""]
            avg_rating = round(sum(ratings) / len(ratings), 1) if ratings else 0
            high_value = sum(
                1 for x in items
                if x["has_website"] == "No" and int(x["review_count"] or 0) >= 20
            )
            summary[city][industry] = {
                "total":          total,
                "no_website":     no_website,
                "no_website_pct": round(no_website / total * 100) if total else 0,
                "avg_rating":     avg_rating,
                "high_value":     high_value,
            }
    return summary


# ── CSV Export ────────────────────────────────────────────

def save_raw_csv(leads: list, path: str):
    if not leads:
        return
    fields = ["city", "area", "industry", "name", "phone", "website",
              "has_website", "rating", "review_count", "address", "place_id"]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(leads)
    print(f"Raw CSV saved: {path}")


# ── Excel Export ──────────────────────────────────────────

def save_excel(leads: list, summary: dict, path: str):
    if not HAS_OPENPYXL:
        print("Skipping Excel — install openpyxl first.")
        return

    wb = openpyxl.Workbook()

    C_DARK   = "1A1A1A"
    C_ACCENT = "C8F135"
    C_WHITE  = "FFFFFF"
    C_LGREY  = "F5F5F5"
    C_RED    = "FF4D4D"
    C_ORANGE = "FF9F43"
    C_GREEN  = "2E7D32"
    C_HEADER = "2D2D2D"

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def font(bold=False, color="000000", size=10):
        return Font(bold=bold, color=color, size=size)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin   = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Overview sheet ────────────────────────────────────
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = "FLORIDA MARKET SCAN — LEADGEN INTELLIGENCE"
    c.font      = Font(bold=True, color=C_ACCENT, size=16)
    c.fill      = fill(C_DARK)
    c.alignment = center()
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value     = f"Generated {datetime.now().strftime('%B %d, %Y')}  ·  9 Cities  ·  17 Industries  ·  5 Areas per City"
    c.font      = Font(color="AAAAAA", size=10)
    c.fill      = fill(C_DARK)
    c.alignment = center()
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 10

    total_biz     = len(leads)
    total_no_site = sum(1 for l in leads if l["has_website"] == "No")
    total_hv      = sum(
        1 for l in leads
        if l["has_website"] == "No" and int(l["review_count"] or 0) >= 20
    )

    stat_data = [
        ("Total Businesses Found",                        f"{total_biz:,}"),
        ("Have No Website",                               f"{total_no_site:,}  ({round(total_no_site/total_biz*100)}%)"),
        ("High-Value Leads\n(No website + 20+ reviews)", f"{total_hv:,}"),
    ]

    col = 1
    for label, value in stat_data:
        for row, val, fnt, bg in [
            (4, label, Font(bold=True, color=C_ACCENT, size=9), C_HEADER),
            (5, value, Font(bold=True, color=C_WHITE, size=14), C_DARK),
        ]:
            c = ws.cell(row=row, column=col, value=val)
            c.font      = fnt
            c.fill      = fill(bg)
            c.alignment = center()
            c.border    = border
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
        col += 2

    ws.row_dimensions[4].height = 28
    ws.row_dimensions[5].height = 36
    ws.row_dimensions[6].height = 16

    hdrs = ["City", "Industry", "Total", "No Website",
            "No Website %", "Avg Rating", "High Value\nLeads", "Opportunity\nScore"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=7, column=ci, value=h)
        c.font      = font(bold=True, color=C_WHITE, size=9)
        c.fill      = fill(C_HEADER)
        c.alignment = center()
        c.border    = border
    ws.row_dimensions[7].height = 30

    all_rows = []
    for city, industries in summary.items():
        for industry, stats in industries.items():
            opp = round((stats["no_website_pct"] * 0.6) + (min(stats["high_value"], 20) * 2))
            all_rows.append((city, industry, stats, opp))
    all_rows.sort(key=lambda x: x[3], reverse=True)

    for row_idx, (city, industry, stats, opp) in enumerate(all_rows, 8):
        pct = stats["no_website_pct"]
        bg  = "FFF3F3" if pct >= 50 else "FFF9F0" if pct >= 30 else C_LGREY
        vals = [city, industry, stats["total"], stats["no_website"],
                f"{pct}%", stats["avg_rating"], stats["high_value"], opp]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=ci, value=val)
            c.fill      = fill(bg)
            c.alignment = center() if ci > 2 else left()
            c.border    = border
            c.font      = Font(size=9)
            if ci == 5:
                if pct >= 50:   c.font = Font(bold=True, color=C_RED,    size=9)
                elif pct >= 30: c.font = Font(bold=True, color=C_ORANGE, size=9)
                else:           c.font = Font(color=C_GREEN, size=9)
            if ci == 8:
                if opp >= 60:   c.font = Font(bold=True, color=C_RED,    size=9)
                elif opp >= 35: c.font = Font(bold=True, color=C_ORANGE, size=9)
        ws.row_dimensions[row_idx].height = 18

    for i, w in enumerate([18, 24, 8, 10, 11, 10, 11, 11], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Per-city sheets ───────────────────────────────────
    for city, industries in summary.items():
        ws2 = wb.create_sheet(title=city[:31])
        ws2.sheet_view.showGridLines = False

        ws2.merge_cells("A1:G1")
        c = ws2["A1"]
        c.value     = f"{city.upper()} — INDUSTRY BREAKDOWN"
        c.font      = Font(bold=True, color=C_ACCENT, size=13)
        c.fill      = fill(C_DARK)
        c.alignment = center()
        ws2.row_dimensions[1].height = 30

        for ci, h in enumerate(["Industry", "Total", "No Website",
                                  "No Website %", "Avg Rating",
                                  "High Value\nLeads", "Opportunity\nScore"], 1):
            c = ws2.cell(row=2, column=ci, value=h)
            c.font      = font(bold=True, color=C_WHITE, size=9)
            c.fill      = fill(C_HEADER)
            c.alignment = center()
            c.border    = border
        ws2.row_dimensions[2].height = 28

        city_rows = sorted(
            [(ind, stats, round((stats["no_website_pct"]*0.6)+(min(stats["high_value"],20)*2)))
             for ind, stats in industries.items()],
            key=lambda x: x[2], reverse=True
        )

        for ri, (ind, stats, opp) in enumerate(city_rows, 3):
            pct = stats["no_website_pct"]
            bg  = "FFF3F3" if pct >= 50 else "FFF9F0" if pct >= 30 else C_LGREY
            for ci, val in enumerate([ind, stats["total"], stats["no_website"],
                                       f"{pct}%", stats["avg_rating"],
                                       stats["high_value"], opp], 1):
                c = ws2.cell(row=ri, column=ci, value=val)
                c.fill      = fill(bg)
                c.alignment = center() if ci > 1 else left()
                c.border    = border
                c.font      = Font(size=9)
                if ci == 4:
                    if pct >= 50:   c.font = Font(bold=True, color=C_RED,    size=9)
                    elif pct >= 30: c.font = Font(bold=True, color=C_ORANGE, size=9)
                    else:           c.font = Font(color=C_GREEN, size=9)
                if ci == 7:
                    if opp >= 60:   c.font = Font(bold=True, color=C_RED,    size=9)
                    elif opp >= 35: c.font = Font(bold=True, color=C_ORANGE, size=9)
            ws2.row_dimensions[ri].height = 18

        for i, w in enumerate([26, 8, 10, 11, 10, 11, 11], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Raw data sheet ────────────────────────────────────
    ws3 = wb.create_sheet(title="Raw Data")
    ws3.sheet_view.showGridLines = False

    raw_hdrs = ["City", "Area", "Industry", "Business Name",
                "Phone", "Website", "Has Website",
                "Rating", "Reviews", "Address"]
    for ci, h in enumerate(raw_hdrs, 1):
        c = ws3.cell(row=1, column=ci, value=h)
        c.font      = font(bold=True, color=C_WHITE, size=9)
        c.fill      = fill(C_HEADER)
        c.alignment = center()
        c.border    = border
    ws3.row_dimensions[1].height = 22

    for ri, lead in enumerate(leads, 2):
        bg = "FFF3F3" if lead["has_website"] == "No" else C_LGREY
        for ci, val in enumerate([
            lead["city"], lead["area"], lead["industry"],
            lead["name"], lead["phone"], lead["website"],
            lead["has_website"], lead["rating"],
            lead["review_count"], lead["address"]
        ], 1):
            c = ws3.cell(row=ri, column=ci, value=val)
            c.fill      = fill(bg)
            c.alignment = left()
            c.border    = border
            c.font      = Font(size=8)
        ws3.row_dimensions[ri].height = 15

    for i, w in enumerate([14, 22, 20, 30, 15, 35, 10, 8, 8, 40], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    wb.save(path)
    print(f"Excel saved: {path}")


# ── Entry Point ───────────────────────────────────────────

if __name__ == "__main__":
    print("LeadGen Florida Market Scanner (Parallel)")
    print("=" * 60)
    total = len(MARKETS) * 5 * len(INDUSTRIES)
    print(f"Cities: {len(MARKETS)}  |  Areas: 5  |  Industries: {len(INDUSTRIES)}")
    print(f"Total queries: {total}  |  Concurrency: {CONCURRENCY}")
    print("=" * 60)

    os.system("pip install openpyxl aiohttp -q")

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True

    start   = time.time()
    leads   = asyncio.run(run_scan_async())
    elapsed = round(time.time() - start)
    summary = build_summary(leads)

    ts         = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path   = f"florida_raw_{ts}.csv"
    excel_path = f"florida_report_{ts}.xlsx"

    save_raw_csv(leads, csv_path)
    save_excel(leads, summary, excel_path)

    print(f"\nCompleted in {elapsed}s ({elapsed//60}m {elapsed%60}s)")
    print(f"Raw CSV  : {csv_path}")
    print(f"Excel    : {excel_path}")

    print("\nTOP 10 OPPORTUNITIES:")
    all_opps = [
        (round((s["no_website_pct"]*0.6)+(min(s["high_value"],20)*2)), city, ind, s)
        for city, inds in summary.items()
        for ind, s in inds.items()
    ]
    all_opps.sort(reverse=True)
    for opp, city, ind, s in all_opps[:10]:
        print(f"  {opp:3d}  {city:<20} {ind:<28} "
              f"{s['no_website_pct']}% no website, "
              f"{s['high_value']} high-value leads")